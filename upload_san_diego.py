"""Upload San Diego restaurant data to Supabase and Pinecone.

Usage:
    python upload_san_diego.py

Requires: openpyxl, httpx, pinecone, openai, python-dotenv
"""

import os
import time
import json
from pathlib import Path

import httpx
import openpyxl
from dotenv import load_dotenv
from openai import OpenAI
from pinecone import Pinecone

load_dotenv()

EXCEL_PATH = Path.home() / "Downloads" / "san_diego_restaurants (1).xlsx"

SUPABASE_URL = os.getenv("NEXT_PUBLIC_SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("NEXT_PUBLIC_SUPABASE_ANON_KEY", "")

PINECONE_API_KEY = os.getenv("PINECONE_API_KEY", "")
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY", "")
OPENROUTER_BASE_URL = os.getenv("OPENROUTER_BASE_URL", "https://openrouter.ai/api/v1")

INDEX_NAME = "food-rescue-menus"
EMBEDDING_MODEL = "google/gemini-embedding-001"
EMBEDDING_DIM = 768
BATCH_SIZE = 50


def read_supabase_sheet(wb):
    ws = wb["Supabase (Relational)"]
    headers = [cell.value for cell in ws[2]]
    rows = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        record = dict(zip(headers, row))
        rows.append(record)
    return rows


def read_pinecone_sheet(wb):
    ws = wb["Pinecone (Vector DB)"]
    headers = [cell.value for cell in ws[2]]
    rows = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        record = dict(zip(headers, row))
        rows.append(record)
    return rows


def upload_to_supabase(rows):
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("⚠  Skipping Supabase: missing NEXT_PUBLIC_SUPABASE_URL or NEXT_PUBLIC_SUPABASE_ANON_KEY")
        return False

    url = f"{SUPABASE_URL}/rest/v1/san_diego_restaurants"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates",
    }

    payloads = []
    for r in rows:
        payloads.append({
            "id": r["id"],
            "name": r["name"],
            "address": r["address"],
            "lat": float(r["lat"]) if r["lat"] else None,
            "lng": float(r["lng"]) if r["lng"] else None,
            "rating": float(r["rating"]) if r["rating"] else None,
            "cuisine": r.get("cuisine", ""),
            "price_level": r.get("price_level", ""),
            "phone": r.get("phone", ""),
            "closing_time": r.get("closing_time", ""),
            "hours_of_operation": r.get("hours_of_operation", ""),
            "menu_items": r.get("menu_items", ""),
        })

    print(f"\n--- Supabase Upload ---")
    print(f"Uploading {len(payloads)} restaurants to {SUPABASE_URL}...")

    for i in range(0, len(payloads), BATCH_SIZE):
        batch = payloads[i:i + BATCH_SIZE]
        resp = httpx.post(url, headers=headers, json=batch, timeout=30)
        if resp.status_code in (200, 201):
            print(f"  Batch {i // BATCH_SIZE + 1}: {len(batch)} rows inserted")
        else:
            print(f"  Batch {i // BATCH_SIZE + 1} FAILED ({resp.status_code}): {resp.text}")
            return False

    print(f"✓ Supabase: {len(payloads)} restaurants uploaded")
    return True


def upload_to_pinecone(rows):
    if not PINECONE_API_KEY:
        print("⚠  Skipping Pinecone: missing PINECONE_API_KEY")
        return False
    if not OPENROUTER_API_KEY:
        print("⚠  Skipping Pinecone: missing OPENROUTER_API_KEY (needed for embeddings)")
        return False

    print(f"\n--- Pinecone Upload ---")

    oai = OpenAI(api_key=OPENROUTER_API_KEY, base_url=OPENROUTER_BASE_URL)
    pc = Pinecone(api_key=PINECONE_API_KEY)

    if pc.has_index(INDEX_NAME):
        desc = pc.describe_index(INDEX_NAME)
        if desc.dimension != EMBEDDING_DIM:
            print(f"  Index exists with dimension {desc.dimension}, need {EMBEDDING_DIM}. Deleting...")
            pc.delete_index(INDEX_NAME)
            time.sleep(5)

    if not pc.has_index(INDEX_NAME):
        print(f"  Creating index '{INDEX_NAME}' (dim={EMBEDDING_DIM})...")
        pc.create_index(
            name=INDEX_NAME,
            dimension=EMBEDDING_DIM,
            metric="cosine",
            spec={"serverless": {"cloud": "aws", "region": "us-east-1"}},
        )
        while not pc.describe_index(INDEX_NAME).status.ready:
            time.sleep(2)
        print(f"  Index '{INDEX_NAME}' is ready")

    index = pc.Index(INDEX_NAME)

    texts = [r["text_for_embedding"] for r in rows]
    vectors = []

    print(f"Generating embeddings for {len(texts)} restaurants...")
    for i in range(0, len(texts), 20):
        batch_texts = texts[i:i + 20]
        resp = oai.embeddings.create(input=batch_texts, model=EMBEDDING_MODEL, dimensions=EMBEDDING_DIM)
        for j, emb_data in enumerate(resp.data):
            row = rows[i + j]
            hours_str = row.get("metadata_hours_of_operation", "")
            hours_list = [h.strip() for h in hours_str.split("|")] if hours_str else []

            vectors.append({
                "id": row["vector_id"],
                "values": emb_data.embedding,
                "metadata": {
                    "restaurant_name": row.get("metadata_name", ""),
                    "address": row.get("metadata_address", ""),
                    "lat": float(row.get("metadata_lat", 0)),
                    "lng": float(row.get("metadata_lng", 0)),
                    "rating": float(row.get("metadata_rating", 0)),
                    "cuisine": row.get("metadata_cuisine", ""),
                    "price_level": row.get("metadata_price_level", ""),
                    "phone": row.get("metadata_phone", ""),
                    "closing_time": row.get("metadata_closing_time", "Unknown"),
                    "hours_of_operation": hours_list,
                },
            })
        print(f"  Embedded {min(i + 20, len(texts))}/{len(texts)}")
        time.sleep(0.5)

    print(f"Upserting {len(vectors)} vectors to '{INDEX_NAME}'...")
    for i in range(0, len(vectors), 100):
        batch = vectors[i:i + 100]
        index.upsert(vectors=batch, namespace="san_diego")
        print(f"  Upserted {min(i + 100, len(vectors))}/{len(vectors)}")

    time.sleep(3)
    stats = index.describe_index_stats()
    print(f"✓ Pinecone: {stats.total_vector_count} total vectors in '{INDEX_NAME}'")
    return True


def main():
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    supabase_rows = read_supabase_sheet(wb)
    pinecone_rows = read_pinecone_sheet(wb)
    print(f"Loaded {len(supabase_rows)} Supabase rows, {len(pinecone_rows)} Pinecone rows")

    upload_to_supabase(supabase_rows)
    upload_to_pinecone(pinecone_rows)

    print("\nDone!")


if __name__ == "__main__":
    main()
